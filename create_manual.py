# -*- coding: utf-8 -*-
"""取扱説明書.xlsx を生成するスクリプト"""
from openpyxl import load_workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

OUTPUT = '取扱説明書.xlsx'
wb = load_workbook(OUTPUT)

# 既存シートを全削除して作り直す
for sn in wb.sheetnames:
    del wb[sn]

# ============================================================
# 共通スタイル
# ============================================================
thin  = Side(style='thin')
med   = Side(style='medium')
BORDER      = Border(left=thin,  right=thin,  top=thin,  bottom=thin)
BORDER_MED  = Border(left=med,   right=med,   top=med,   bottom=med)
BORDER_BOT  = Border(bottom=med)

def font(name='メイリオ', size=9, bold=False, color='000000', italic=False):
    return Font(name=name, size=size, bold=bold, color=color, italic=italic)

def fill(color):
    return PatternFill('solid', fgColor=color)

def cell(ws, r, c, v, fnt=None, fl=None, align=None, border=None, span=None):
    if span:
        ws.merge_cells(f'{get_column_letter(c)}{r}:{get_column_letter(c+span-1)}{r}')
    ce = ws.cell(row=r, column=c, value=v)
    if fnt:    ce.font      = fnt
    if fl:     ce.fill      = fl
    if align:  ce.alignment = align
    if border: ce.border    = border
    return ce

C  = Alignment(horizontal='center', vertical='center', wrap_text=True)
L  = Alignment(horizontal='left',   vertical='center', wrap_text=True)
LT = Alignment(horizontal='left',   vertical='top',    wrap_text=True)

# カラーパレット
C_NAVY   = '1F4E79'
C_BLUE   = '2E75B6'
C_LBLUE  = '5B9BD5'
C_SKYBLUE= 'BDD7EE'
C_PALE   = 'EBF3FB'
C_WHITE  = 'FFFFFF'
C_GRAY   = 'F2F2F2'
C_LGRAY  = 'D9D9D9'
C_GREEN  = '375623'
C_LGREEN = 'C6EFCE'
C_ORANGE = 'C55A11'
C_LORANGE= 'FCE4D6'
C_YELLOW = 'FFEB9C'
C_RED    = '9C0006'
C_LRED   = 'FFC7CE'

# ============================================================
# ページ1: 表紙
# ============================================================
ws = wb.create_sheet('表紙')
ws.sheet_view.showGridLines = False

for i, w in enumerate([2,20,20,20,20,20,20,2], 1):
    ws.column_dimensions[get_column_letter(i)].width = w
for r in range(1, 50):
    ws.row_dimensions[r].height = 18

# 全体背景
for r in range(1, 50):
    for c in range(1, 9):
        ws.cell(row=r, column=c).fill = fill('F5F8FC')

# タイトルブロック
for r in range(6, 14):
    for c in range(2, 8):
        ws.cell(row=r, column=c).fill = fill(C_NAVY)

ws.merge_cells('B6:G13')
c = ws.cell(row=6, column=2,
    value='混載パレタイズ積付計算システム\n取 扱 説 明 書')
c.font      = Font(name='メイリオ', size=22, bold=True, color=C_WHITE)
c.fill      = fill(C_NAVY)
c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
ws.row_dimensions[6].height = 30

# サブタイトル
ws.merge_cells('B14:G14')
c = ws.cell(row=14, column=2, value='Palletizing Optimization System  ver 1.1.0')
c.font      = font(size=11, color=C_NAVY)
c.fill      = fill(C_SKYBLUE)
c.alignment = C

# 情報ボックス
info = [
    ('システム名',   '混載パレタイズ積付計算システム'),
    ('バージョン',   'ver 1.1.0'),
    ('URL',          'https://mixed-palletetize.onrender.com'),
    ('作成日',       '2026年3月'),
]
for i, (lbl, val) in enumerate(info):
    r = 17 + i * 2
    ws.merge_cells(f'C{r}:C{r}')
    ws.merge_cells(f'D{r}:G{r}')
    cell(ws, r, 3, lbl, fnt=font(bold=True, color=C_WHITE),
         fl=fill(C_BLUE), align=C, border=BORDER)
    cell(ws, r, 4, val, fnt=font(size=10),
         fl=fill(C_PALE), align=L, border=BORDER, span=4)
    ws.row_dimensions[r].height = 22

# 注意書き
ws.merge_cells('B30:G32')
c = ws.cell(row=30, column=2,
    value='本書はシステムの基本的な操作方法を説明するものです。\n'
          'システムの仕様変更により、一部の画面・機能が変更される場合があります。')
c.font      = font(size=9, italic=True, color='595959')
c.alignment = LT

# ============================================================
# ページ2: 目次
# ============================================================
ws = wb.create_sheet('目次')
ws.sheet_view.showGridLines = False

for i, w in enumerate([2,6,30,40,12,2], 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# タイトル
ws.merge_cells('B2:E2')
cell(ws, 2, 2, '目　次', fnt=font(size=16, bold=True, color=C_WHITE),
     fl=fill(C_NAVY), align=C, span=4)
ws.row_dimensions[2].height = 30

# 目次項目
toc = [
    ('1', 'システム概要',           'システムの目的・特徴・動作環境'),
    ('2', 'アクセス方法・ログイン', 'URLアクセス・Basic認証'),
    ('3', '基本操作フロー',         '計算実行までの流れ（5ステップ）'),
    ('4', 'ケース情報入力',         'CSVインポート・手動入力・詳細属性'),
    ('5', 'パレット・積付設定',     'パレット寸法・供給制約・積付ルール・積付制約'),
    ('6', '評価設定',               'プリセット・評価重み・解探索モード'),
    ('7', '動作モード',             '理想最適・現実制約・ロボ実行モードの違い'),
    ('8', '積付結果の見方',         'KPI・3Dビュー・平面図・サイドビュー・アニメーション'),
    ('9', 'CSVエクスポート',        '結果のCSVダウンロード'),
    ('10','注意事項・制限',         '入力制限・スリープ仕様・推奨ブラウザ'),
]
ws.row_dimensions[3].height = 6
for i, (no, title, desc) in enumerate(toc):
    r = 4 + i * 2
    cell(ws, r, 2, no,    fnt=font(bold=True, color=C_WHITE), fl=fill(C_LBLUE), align=C, border=BORDER)
    cell(ws, r, 3, title, fnt=font(bold=True, size=10),       fl=fill(C_PALE),  align=L, border=BORDER)
    cell(ws, r, 4, desc,  fnt=font(size=9, color='595959'),   fl=fill(C_GRAY),  align=L, border=BORDER)
    cell(ws, r, 5, f'シート「{no}.{title[:6]}」', fnt=font(size=8, color=C_BLUE), align=C, border=BORDER)
    ws.row_dimensions[r].height = 20
    ws.row_dimensions[r+1].height = 4

# ============================================================
# 共通ヘッダー描画関数
# ============================================================
def draw_sheet_header(ws, no, title, desc):
    ws.sheet_view.showGridLines = False
    ws.merge_cells('B1:G1')
    c = ws.cell(row=1, column=1)
    ws.merge_cells('A1:A1')

    ws.merge_cells('B1:G1')
    ce = ws.cell(row=1, column=2, value=f'{no}. {title}')
    ce.font = Font(name='メイリオ', size=14, bold=True, color=C_WHITE)
    ce.fill = fill(C_NAVY)
    ce.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 28

    ws.merge_cells('B2:G2')
    ce = ws.cell(row=1, column=2)
    ce2 = ws.cell(row=2, column=2, value=desc)
    ce2.font = Font(name='メイリオ', size=9, italic=True, color='595959')
    ce2.fill = fill(C_SKYBLUE)
    ce2.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[2].height = 16

def section(ws, r, title, col_span=6):
    ws.merge_cells(f'B{r}:{get_column_letter(1+col_span)}{r}')
    ce = ws.cell(row=r, column=2, value=f'■ {title}')
    ce.font = Font(name='メイリオ', size=10, bold=True, color=C_WHITE)
    ce.fill = fill(C_BLUE)
    ce.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[r].height = 20

def note_cell(ws, r, text, col_span=6):
    ws.merge_cells(f'B{r}:{get_column_letter(1+col_span)}{r}')
    ce = ws.cell(row=r, column=2, value=text)
    ce.font = Font(name='メイリオ', size=9, italic=True, color='595959')
    ce.fill = fill(C_YELLOW)
    ce.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws.row_dimensions[r].height = 32

# ============================================================
# ページ3: システム概要
# ============================================================
ws = wb.create_sheet('1.システム概要')
for i, w in enumerate([2,20,18,18,18,18,2], 1):
    ws.column_dimensions[get_column_letter(i)].width = w
draw_sheet_header(ws, 1, 'システム概要', 'システムの目的・主要機能・動作環境について説明します')

r = 4
section(ws, r, 'システムの目的')
r += 1
ws.merge_cells(f'B{r}:G{r}')
ce = ws.cell(row=r, column=2,
    value='本システムは、異なるサイズ・重量の複数品種の荷物（ケース）を、1台または複数台のパレットに効率よく積み付けるための'
          '計算・可視化ツールです。現場の供給制約（FIFO・バッファ）や積付ルールを考慮した最適な積付パターンを自動計算します。')
ce.font = Font(name='メイリオ', size=9)
ce.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
ws.row_dimensions[r].height = 45
r += 2

section(ws, r, '主要機能')
r += 1
features = [
    ('① ケース情報入力',    'CSV/Excelファイルのインポート、または手動入力で最大30品種のケース情報を登録'),
    ('② 積付計算',          'ビームサーチアルゴリズムによる高精度な積付パターンの自動計算'),
    ('③ 3D可視化',          'Three.jsによるインタラクティブな等角3Dビューで積付結果を確認'),
    ('④ 平面図・側面図',    '真上・側面からの2D図による積付状態の確認'),
    ('⑤ アニメーション再生','ケースが積まれていく順序をアニメーションで確認'),
    ('⑥ 結果CSV出力',       '積付結果をCSVファイルとしてダウンロード'),
    ('⑦ 理想解との比較',    '制約なし理想解と実際の計算結果の差分を自動分析'),
]
for i, (f_title, f_desc) in enumerate(features):
    fl = fill(C_PALE) if i % 2 == 0 else fill(C_GRAY)
    cell(ws, r, 2, f_title, fnt=font(bold=True, size=9), fl=fl, align=L, border=BORDER)
    ws.merge_cells(f'C{r}:G{r}')
    cell(ws, r, 3, f_desc,  fnt=font(size=9),            fl=fl, align=L, border=BORDER, span=5)
    ws.row_dimensions[r].height = 18
    r += 1
r += 1

section(ws, r, '動作環境')
r += 1
envs = [
    ('推奨ブラウザ', 'Google Chrome / Microsoft Edge（最新版）'),
    ('インターネット接続', '必要（オンラインサービス）'),
    ('URL', 'https://mixed-palletetize.onrender.com'),
    ('認証', 'Basic認証（ID・パスワードが必要）'),
    ('注意', 'Internet Explorer は非対応。Three.js使用のためWebGL対応ブラウザが必要。'),
]
for i, (lbl, val) in enumerate(envs):
    fl = fill(C_PALE) if i % 2 == 0 else fill(C_GRAY)
    cell(ws, r, 2, lbl, fnt=font(bold=True, size=9), fl=fl, align=C, border=BORDER)
    ws.merge_cells(f'C{r}:G{r}')
    cell(ws, r, 3, val, fnt=font(size=9),            fl=fl, align=L, border=BORDER, span=5)
    ws.row_dimensions[r].height = 18
    r += 1

# ============================================================
# ページ4: アクセス方法
# ============================================================
ws = wb.create_sheet('2.アクセス・ログイン')
for i, w in enumerate([2,8,25,35,22,2], 1):
    ws.column_dimensions[get_column_letter(i)].width = w
draw_sheet_header(ws, 2, 'アクセス方法・ログイン', 'システムへのアクセス手順を説明します')

r = 4
section(ws, r, 'アクセス手順')
r += 1
steps = [
    ('Step 1', 'ブラウザを起動',
     'Google Chrome または Microsoft Edge を起動します。'),
    ('Step 2', 'URLにアクセス',
     'アドレスバーに以下を入力してEnterキーを押します。\nhttps://mixed-palletetize.onrender.com'),
    ('Step 3', 'ログイン',
     '「認証が必要です」ダイアログが表示されます。\nシステム管理者から提供されたIDとパスワードを入力して「ログイン」をクリックします。'),
    ('Step 4', '画面表示',
     'システムのメイン画面が表示されれば準備完了です。'),
]
for i, (no, title, desc) in enumerate(steps):
    fl_h = fill(C_LBLUE) if i % 2 == 0 else fill(C_BLUE)
    cell(ws, r, 2, no,    fnt=font(bold=True, color=C_WHITE, size=10), fl=fl_h,      align=C, border=BORDER)
    cell(ws, r, 3, title, fnt=font(bold=True, size=10),                fl=fill(C_PALE), align=L, border=BORDER)
    ws.merge_cells(f'D{r}:E{r}')
    cell(ws, r, 4, desc,  fnt=font(size=9), fl=fill(C_GRAY), align=LT, border=BORDER, span=2)
    ws.row_dimensions[r].height = 40
    r += 1
r += 1

note_cell(ws, r, '⚠ 注意：アクセスから15分間操作がない場合、サーバーがスリープ状態になります。\n'
                  '再アクセス時は画面が表示されるまで30秒〜1分程度かかる場合があります。これは正常な動作です。')
r += 2

section(ws, r, '画面構成')
r += 1
areas = [
    ('ヘッダー部',   '画面上部。システム名・パレット寸法等のスペック情報を表示。'),
    ('動作モードバー', 'ヘッダー直下。理想最適・現実制約・ロボ実行の3モードを切り替えます。'),
    ('タブ①',       'ケース情報入力タブ。積み付けるケースの情報を登録します。'),
    ('タブ②',       'パレット・積付設定タブ。パレット寸法・供給制約・積付ルールを設定します。'),
    ('タブ③',       '評価設定タブ。最適化の評価重みや解探索モードを設定します。'),
    ('タブ④',       '積付結果タブ。計算結果のKPI・3Dビュー・図面を表示します。'),
]
for i, (area, desc) in enumerate(areas):
    fl = fill(C_PALE) if i % 2 == 0 else fill(C_GRAY)
    cell(ws, r, 2, area, fnt=font(bold=True, size=9), fl=fl, align=C, border=BORDER)
    ws.merge_cells(f'C{r}:E{r}')
    cell(ws, r, 3, desc, fnt=font(size=9), fl=fl, align=L, border=BORDER, span=3)
    ws.row_dimensions[r].height = 18
    r += 1

# ============================================================
# ページ5: 基本操作フロー
# ============================================================
ws = wb.create_sheet('3.基本操作フロー')
for i, w in enumerate([2,10,28,42,10,2], 1):
    ws.column_dimensions[get_column_letter(i)].width = w
draw_sheet_header(ws, 3, '基本操作フロー', '計算実行から結果確認までの基本的な流れです')

r = 4
section(ws, r, '操作の流れ（5ステップ）')
r += 1
flows = [
    ('STEP 1', 'ケース情報を入力する',
     'タブ①「ケース情報入力」でCSVファイルをインポートするか、手動でケース情報を入力します。',
     '品番・品名・サイズ・重量・数量が必須項目です。'),
    ('STEP 2', 'パレット・積付設定を確認する',
     'タブ②「パレット・積付設定」でパレット寸法・供給モード・積付ルールを確認・変更します。',
     '通常はデフォルト設定のまま使用できます。'),
    ('STEP 3', '評価設定を確認する',
     'タブ③「評価設定」で最適化プリセットや評価重みを設定します。',
     '「標準」プリセットで多くのケースに対応できます。'),
    ('STEP 4', '計算を実行する',
     'タブ①下部の「▶ 計算実行」ボタンをクリックします。',
     '計算中は「計算中...」のトーストが表示されます。数秒〜数十秒で完了します。'),
    ('STEP 5', '結果を確認する',
     'タブ④「積付結果」に自動で切り替わります。KPI・3Dビュー・平面図を確認してください。',
     '結果はCSVでダウンロードすることもできます。'),
]
colors = [C_NAVY, C_BLUE, C_LBLUE, '2E86C1', '5DADE2']
for i, (step, title, desc, hint) in enumerate(flows):
    cell(ws, r, 2, step,  fnt=font(bold=True, color=C_WHITE, size=11), fl=fill(colors[i]), align=C, border=BORDER)
    cell(ws, r, 3, title, fnt=font(bold=True, size=11),                fl=fill(C_PALE),    align=L, border=BORDER)
    cell(ws, r, 4, desc,  fnt=font(size=9),                            fl=fill(C_GRAY),    align=LT, border=BORDER)
    cell(ws, r, 5, '💡',  fnt=font(size=10),                           fl=fill(C_YELLOW),  align=C, border=BORDER)
    ws.row_dimensions[r].height = 14
    r += 1
    ws.merge_cells(f'C{r}:D{r}')
    cell(ws, r, 3, f'→ {hint}', fnt=font(size=8, color='595959', italic=True),
         fl=fill('FAFAFA'), align=LT, border=BORDER, span=2)
    cell(ws, r, 5, '',           fl=fill(C_YELLOW), border=BORDER)
    ws.row_dimensions[r].height = 16
    r += 1
r += 1

note_cell(ws, r,
    '💡 ポイント：設定変更後は必ず「▶ 計算実行」を再度クリックしてください。設定を変えても結果は自動更新されません。')

# ============================================================
# ページ6: ケース情報入力
# ============================================================
ws = wb.create_sheet('4.ケース情報入力')
for i, w in enumerate([2,14,20,42,14,2], 1):
    ws.column_dimensions[get_column_letter(i)].width = w
draw_sheet_header(ws, 4, 'ケース情報入力', 'タブ①の操作方法・入力項目の説明です')

r = 4
section(ws, r, 'CSVファイルからのインポート（推奨）')
r += 1
ws.merge_cells(f'B{r}:E{r}')
ce = ws.cell(row=r, column=2,
    value='「📂 CSV / Excel 読込」ボタンをクリックし、CSVまたはExcelファイルを選択します。')
ce.font = Font(name='メイリオ', size=9); ce.alignment = L; ce.fill = fill(C_PALE)
ws.row_dimensions[r].height = 18
r += 1

# CSVフォーマット
cell(ws, r, 2, 'CSVフォーマット', fnt=font(bold=True, size=9, color=C_WHITE), fl=fill(C_LBLUE), align=C, border=BORDER, span=4)
ws.row_dimensions[r].height = 16
r += 1
headers_csv = ['列番号', '項目名', '必須/任意', '説明']
for ci, h in enumerate(headers_csv, 2):
    cell(ws, r, ci, h, fnt=font(bold=True, size=8, color=C_WHITE), fl=fill(C_BLUE), align=C, border=BORDER)
ws.row_dimensions[r].height = 16
r += 1
csv_cols = [
    ('1列目', '品番（SKU ID）', '必須', '荷物を識別する番号・コード'),
    ('2列目', '品名',           '必須', '荷物の名称'),
    ('3列目', '長さ L (mm)',    '必須', 'X方向のサイズ（ミリメートル）'),
    ('4列目', '幅 W (mm)',      '必須', 'Y方向のサイズ（ミリメートル）'),
    ('5列目', '高さ H (mm)',    '必須', 'Z方向のサイズ（ミリメートル）'),
    ('6列目', '重量 (kg)',      '必須', '1ケースあたりの重量'),
    ('7列目', '数量',           '必須', '積み付けるケース数'),
    ('8列目', 'グループ',       '任意', '同品種集約用のグループ名'),
    ('9列目', 'fragile',        '任意', '壊れやすい場合は "1" または "true"'),
    ('10列目','no_flip',         '任意', '天地無用の場合は "1" または "true"'),
    ('11列目','max_top_load',    '任意', '上面許容荷重 (kg)。0=制限なし'),
    ('12列目','温度帯',          '任意', '"常温" / "冷蔵" / "冷凍"'),
]
for i, (col, name, req, desc) in enumerate(csv_cols):
    fl = fill(C_PALE) if i % 2 == 0 else fill(C_GRAY)
    cell(ws, r, 2, col,  fnt=font(size=8), fl=fl, align=C, border=BORDER)
    cell(ws, r, 3, name, fnt=font(size=8), fl=fl, align=L, border=BORDER)
    req_fl = fill(C_LRED) if req == '必須' else fill(C_LGREEN)
    req_fn = font(size=8, bold=(req=='必須'), color=(C_RED if req=='必須' else C_GREEN))
    cell(ws, r, 4, req,  fnt=req_fn,       fl=req_fl, align=C, border=BORDER)
    cell(ws, r, 5, desc, fnt=font(size=8), fl=fl, align=L, border=BORDER)
    ws.row_dimensions[r].height = 16
    r += 1
r += 1

section(ws, r, '手動入力')
r += 1
ws.merge_cells(f'B{r}:E{r}')
ce = ws.cell(row=r, column=2,
    value='「＋ ケース追加」ボタンで行を追加し、各セルに直接入力します。最大30品種まで登録できます。')
ce.font = Font(name='メイリオ', size=9); ce.alignment = L; ce.fill = fill(C_PALE)
ws.row_dimensions[r].height = 18
r += 1

ws.merge_cells(f'B{r}:E{r}')
ce = ws.cell(row=r, column=2,
    value='「詳細」ボタン（各行の右端）をクリックすると、fragile・天地無用・上面許容荷重・温度帯などの詳細属性を設定できます。')
ce.font = Font(name='メイリオ', size=9); ce.alignment = L; ce.fill = fill(C_GRAY)
ws.row_dimensions[r].height = 18
r += 2

note_cell(ws, r,
    '⚠ 入力制限：サイズは L:250–500mm / W:150–500mm / H:95–350mm の範囲が推奨です。\n'
    '範囲外の値を入力すると入力欄が赤くハイライトされます。登録前にエラーを修正してください。')

# ============================================================
# ページ7: パレット・積付設定
# ============================================================
ws = wb.create_sheet('5.パレット積付設定')
for i, w in enumerate([2,16,22,44,8,2], 1):
    ws.column_dimensions[get_column_letter(i)].width = w
draw_sheet_header(ws, 5, 'パレット・積付設定', 'タブ②の各設定項目の説明です')

r = 4
section(ws, r, 'A. パレット寸法')
r += 1
pallet_items = [
    ('パレット長さ',     '1100mm', '積み付け面のX方向サイズ'),
    ('パレット幅',       '1100mm', '積み付け面のY方向サイズ'),
    ('パレット高さ',     '150mm',  'パレット台自体の高さ（積み付けには含まない）'),
    ('最大積付高さ',     '1800mm', 'パレット台含む全体の最大高さ制限'),
    ('有効積付高さ',     '1650mm', '実際にケースを積める高さ（最大積付高さ − パレット高さ、自動計算）'),
    ('オーバーハング許容', '25mm', 'ケースがパレット端からはみ出せる最大量'),
]
cell(ws, r, 2, '項目', fnt=font(bold=True, color=C_WHITE), fl=fill(C_BLUE), align=C, border=BORDER)
cell(ws, r, 3, 'デフォルト値', fnt=font(bold=True, color=C_WHITE), fl=fill(C_BLUE), align=C, border=BORDER)
cell(ws, r, 4, '説明', fnt=font(bold=True, color=C_WHITE), fl=fill(C_BLUE), align=L, border=BORDER)
ws.row_dimensions[r].height = 16; r += 1
for i, (name, default, desc) in enumerate(pallet_items):
    fl = fill(C_PALE) if i % 2 == 0 else fill(C_GRAY)
    cell(ws, r, 2, name,    fnt=font(size=9, bold=True), fl=fl, align=L, border=BORDER)
    cell(ws, r, 3, default, fnt=font(size=9),            fl=fl, align=C, border=BORDER)
    cell(ws, r, 4, desc,    fnt=font(size=9),            fl=fl, align=L, border=BORDER)
    ws.row_dimensions[r].height = 16; r += 1
r += 1

section(ws, r, 'B. ケース供給制約（供給モード）')
r += 1
supply_modes = [
    ('🔀 順番自由',    'free',   '供給順を自由に最適化。理論上最も効率が高い。現場のコンベア順番制約がない場合に使用。'),
    ('➡ FIFO制約あり', 'fifo',  '上流からの供給順を維持して積み付け。現場のコンベアライン順序に従う場合に使用（デフォルト）。'),
    ('📦 バッファあり', 'buffer','N件先読みして最適なケースを選択。FIFOと順番自由の中間的なモード。'),
]
cell(ws, r, 2, 'モード名', fnt=font(bold=True, color=C_WHITE), fl=fill(C_BLUE), align=C, border=BORDER)
cell(ws, r, 3, 'コード',   fnt=font(bold=True, color=C_WHITE), fl=fill(C_BLUE), align=C, border=BORDER)
cell(ws, r, 4, '説明',     fnt=font(bold=True, color=C_WHITE), fl=fill(C_BLUE), align=L, border=BORDER)
ws.row_dimensions[r].height = 16; r += 1
for i, (name, code, desc) in enumerate(supply_modes):
    fl = fill(C_PALE) if i % 2 == 0 else fill(C_GRAY)
    cell(ws, r, 2, name, fnt=font(size=9, bold=True), fl=fl, align=L, border=BORDER)
    cell(ws, r, 3, code, fnt=font(size=8),            fl=fl, align=C, border=BORDER)
    cell(ws, r, 4, desc, fnt=font(size=9),            fl=fl, align=LT, border=BORDER)
    ws.row_dimensions[r].height = 30; r += 1
r += 1

section(ws, r, 'D. 積付制約・品質条件（主要チェックボックス）')
r += 1
constraints = [
    ('Fragile品の上載禁止',  'OFF', 'fragile=trueのケースの上には他のケースを積まない'),
    ('重い物を下段優先',      'ON',  '重量の大きいケースを優先的に下段に配置する'),
    ('パレット中心寄せ優先',  'ON',  '重心がパレット中央に来るよう配置を優先する'),
    ('外周優先配置',          'OFF', 'パレットの外壁沿いから配置を開始する'),
    ('段積み優先',            'OFF', '隙間充填より段をそろえることを優先する'),
    ('オーバーハング禁止',    'OFF', 'ケースがパレット端からはみ出ることを完全禁止'),
    ('温度帯混載禁止',        'OFF', '常温・冷蔵・冷凍の異なる温度帯のケースを同一パレットに混載しない'),
]
cell(ws, r, 2, 'ルール名',   fnt=font(bold=True, color=C_WHITE), fl=fill(C_BLUE), align=C, border=BORDER)
cell(ws, r, 3, 'デフォルト', fnt=font(bold=True, color=C_WHITE), fl=fill(C_BLUE), align=C, border=BORDER)
cell(ws, r, 4, '説明',       fnt=font(bold=True, color=C_WHITE), fl=fill(C_BLUE), align=L, border=BORDER)
ws.row_dimensions[r].height = 16; r += 1
for i, (name, default, desc) in enumerate(constraints):
    fl = fill(C_PALE) if i % 2 == 0 else fill(C_GRAY)
    d_fl = fill(C_LGREEN) if default == 'ON' else fill(C_LGRAY)
    d_fn = font(size=9, bold=True, color=C_GREEN if default == 'ON' else '595959')
    cell(ws, r, 2, name,    fnt=font(size=9, bold=True), fl=fl,   align=L, border=BORDER)
    cell(ws, r, 3, default, fnt=d_fn,                    fl=d_fl, align=C, border=BORDER)
    cell(ws, r, 4, desc,    fnt=font(size=9),             fl=fl,   align=LT, border=BORDER)
    ws.row_dimensions[r].height = 20; r += 1

# ============================================================
# ページ8: 評価設定
# ============================================================
ws = wb.create_sheet('6.評価設定')
for i, w in enumerate([2,18,18,42,12,2], 1):
    ws.column_dimensions[get_column_letter(i)].width = w
draw_sheet_header(ws, 6, '評価設定', 'タブ③の最適化プリセット・評価重み・解探索モードの説明です')

r = 4
section(ws, r, '最適化プリセット')
r += 1
presets = [
    ('⚖ 標準',         'standard',  'デフォルト設定。積載率・安定性をバランスよく最適化します（多くの場合はこれで十分）。'),
    ('🛡 安定性優先',   'stability', '支持率・重心安定性を最大化。輸送中の荷崩れリスクを最小化したい場合に使用。'),
    ('📊 積載率優先',   'loadRate',  '体積効率を最大化。パレット数を減らしコスト削減を優先したい場合に使用。'),
    ('🏭 現場再現優先', 'field',     'FIFO・重量順・温度帯制約を重視。現場の実運用ルールに最も忠実な結果を得たい場合。'),
]
cell(ws, r, 2, 'プリセット名', fnt=font(bold=True, color=C_WHITE), fl=fill(C_BLUE), align=C, border=BORDER)
cell(ws, r, 3, 'コード',       fnt=font(bold=True, color=C_WHITE), fl=fill(C_BLUE), align=C, border=BORDER)
cell(ws, r, 4, '特徴・使いどころ', fnt=font(bold=True, color=C_WHITE), fl=fill(C_BLUE), align=L, border=BORDER)
ws.row_dimensions[r].height = 16; r += 1
for i, (name, code, desc) in enumerate(presets):
    fl = fill(C_PALE) if i % 2 == 0 else fill(C_GRAY)
    cell(ws, r, 2, name, fnt=font(size=9, bold=True), fl=fl, align=L, border=BORDER)
    cell(ws, r, 3, code, fnt=font(size=8),            fl=fl, align=C, border=BORDER)
    cell(ws, r, 4, desc, fnt=font(size=9),            fl=fl, align=LT, border=BORDER)
    ws.row_dimensions[r].height = 26; r += 1
r += 1

section(ws, r, '評価重みスライダー')
r += 1
weights = [
    ('積載率重視',    '50%', 'パレット内の体積効率を重視。高いほど隙間なく詰め込もうとする。'),
    ('安定性重視',    '50%', '荷物の支持率・重心安定性を重視。高いほど崩れにくい積み方を優先。'),
    ('高さ抑制重視',  '30%', '積付高さを低く抑えることを重視。高いほど低い積み方を優先。'),
    ('同品種集約重視','30%', '同じ品番のケースを近い位置に集めることを重視。'),
]
cell(ws, r, 2, '評価項目',       fnt=font(bold=True, color=C_WHITE), fl=fill(C_BLUE), align=C, border=BORDER)
cell(ws, r, 3, 'デフォルト値',   fnt=font(bold=True, color=C_WHITE), fl=fill(C_BLUE), align=C, border=BORDER)
cell(ws, r, 4, '説明',           fnt=font(bold=True, color=C_WHITE), fl=fill(C_BLUE), align=L, border=BORDER)
ws.row_dimensions[r].height = 16; r += 1
for i, (name, default, desc) in enumerate(weights):
    fl = fill(C_PALE) if i % 2 == 0 else fill(C_GRAY)
    cell(ws, r, 2, name,    fnt=font(size=9, bold=True), fl=fl, align=L, border=BORDER)
    cell(ws, r, 3, default, fnt=font(size=9),            fl=fl, align=C, border=BORDER)
    cell(ws, r, 4, desc,    fnt=font(size=9),            fl=fl, align=LT, border=BORDER)
    ws.row_dimensions[r].height = 20; r += 1
r += 1

section(ws, r, '解探索モード（ビームサーチ幅）')
r += 1
search_modes = [
    ('高速（貪欲法）',          '1', '最も速い。各ステップで最善の1手のみを探索。精度はやや低い。'),
    ('標準（ビームサーチ width=3）', '3', '速度と精度のバランス型。'),
    ('高精度（ビームサーチ width=5）', '5', 'デフォルト。最も精度が高いが計算時間が増える。推奨設定。'),
]
cell(ws, r, 2, 'モード',     fnt=font(bold=True, color=C_WHITE), fl=fill(C_BLUE), align=C, border=BORDER)
cell(ws, r, 3, 'ビーム幅',   fnt=font(bold=True, color=C_WHITE), fl=fill(C_BLUE), align=C, border=BORDER)
cell(ws, r, 4, '説明',       fnt=font(bold=True, color=C_WHITE), fl=fill(C_BLUE), align=L, border=BORDER)
ws.row_dimensions[r].height = 16; r += 1
for i, (name, width, desc) in enumerate(search_modes):
    fl = fill(C_PALE) if i % 2 == 0 else fill(C_GRAY)
    hl = fill(C_LGREEN) if i == 2 else fl
    cell(ws, r, 2, name,  fnt=font(size=9, bold=(i==2)), fl=hl, align=L, border=BORDER)
    cell(ws, r, 3, width, fnt=font(size=9),              fl=hl, align=C, border=BORDER)
    cell(ws, r, 4, desc,  fnt=font(size=9),              fl=hl, align=LT, border=BORDER)
    ws.row_dimensions[r].height = 20; r += 1

# ============================================================
# ページ9: 動作モード
# ============================================================
ws = wb.create_sheet('7.動作モード')
for i, w in enumerate([2,16,16,56,2], 1):
    ws.column_dimensions[get_column_letter(i)].width = w
draw_sheet_header(ws, 7, '動作モード', '画面上部の3種類の動作モードの違いを説明します')

r = 4
section(ws, r, '動作モードの概要', 5)
r += 1
modes = [
    ('🌟 理想最適モード', 'ideal',
     '【制約】パレット範囲・衝突のみ（最小限）\n'
     '【用途】理論上の最大積載率を確認したいとき。現場制約を一切考慮しない理想解を計算。\n'
     '【注意】FIFO制約と組み合わせると矛盾するため、供給モードは「順番自由」推奨。'),
    ('🏭 現実制約モード', 'real',
     '【制約】全制約を適用（支持率・重量順・fragile・温度帯・オーバーハング等）\n'
     '【用途】現場での実際の積み付けを想定した計算。デフォルト設定。\n'
     '【備考】理想解との差分（制約による効率ロス）も自動で計算・表示されます。'),
    ('🤖 ロボ実行モード', 'robot',
     '【制約】現実制約モードと同等（現時点では同じ計算結果）\n'
     '【用途】ロボットアーム・自動化設備での実行を想定したモード。\n'
     '【将来】把持姿勢・アーム可動域・動作速度制約の追加実装を予定。'),
]
mode_colors = [C_NAVY, 'C55A11', '2471A3']
for i, (name, code, desc) in enumerate(modes):
    cell(ws, r, 2, name, fnt=font(bold=True, color=C_WHITE, size=11),
         fl=fill(mode_colors[i]), align=C, border=BORDER)
    cell(ws, r, 3, f'exec_mode\n= "{code}"', fnt=font(size=8), fl=fill(C_LGRAY), align=C, border=BORDER)
    cell(ws, r, 4, desc, fnt=font(size=9), fl=fill(C_PALE if i%2==0 else C_GRAY), align=LT, border=BORDER)
    ws.row_dimensions[r].height = 55; r += 1
r += 1

note_cell(ws, r,
    '💡 推奨：通常は「🏭 現実制約モード」を使用してください。\n'
    '理想最適モードとの差分を確認することで、現場制約による効率ロスを定量的に把握できます。',
    col_span=4)

# ============================================================
# ページ10: 積付結果の見方
# ============================================================
ws = wb.create_sheet('8.積付結果の見方')
for i, w in enumerate([2,16,18,46,10,2], 1):
    ws.column_dimensions[get_column_letter(i)].width = w
draw_sheet_header(ws, 8, '積付結果の見方', 'タブ④に表示される各KPI・ビュー・図面の説明です')

r = 4
section(ws, r, 'KPI（主要指標）')
r += 1
kpis = [
    ('体積効率',    '全ケース体積 ÷ パレット1枚分の体積 × 100 (%)',
     '高いほど効率よく積まれている。100%超の場合は複数パレット使用。'),
    ('最大高さ',    '積み付けられたケースの最上面の高さ (mm)',
     '有効積付高さ（デフォルト1650mm）以下であれば制約OK。'),
    ('総重量',      '積み付けられた全ケースの合計重量 (kg)',
     '最大積載重量（デフォルト1000kg）以下であれば制約OK。'),
    ('パレット数',  '使用したパレットの総枚数',
     '1枚に収まれば理想。2枚以上の場合は効率や制約を見直す。'),
    ('安定性スコア','積付の安定性を0〜100で評価',
     '高いほど荷崩れしにくい積み方。70以上が推奨。'),
    ('配置数',      '実際に配置できたケース数 / 総ケース数',
     '全数配置できていれば「配置数 = 総ケース数」。'),
]
cell(ws, r, 2, 'KPI名',     fnt=font(bold=True, color=C_WHITE), fl=fill(C_BLUE), align=C, border=BORDER)
cell(ws, r, 3, '計算式・単位', fnt=font(bold=True, color=C_WHITE), fl=fill(C_BLUE), align=L, border=BORDER)
cell(ws, r, 4, '見方・目安', fnt=font(bold=True, color=C_WHITE), fl=fill(C_BLUE), align=L, border=BORDER)
ws.row_dimensions[r].height = 16; r += 1
for i, (name, formula, tips) in enumerate(kpis):
    fl = fill(C_PALE) if i % 2 == 0 else fill(C_GRAY)
    cell(ws, r, 2, name,    fnt=font(size=9, bold=True), fl=fl, align=L, border=BORDER)
    cell(ws, r, 3, formula, fnt=font(size=8),            fl=fl, align=LT, border=BORDER)
    cell(ws, r, 4, tips,    fnt=font(size=9),            fl=fl, align=LT, border=BORDER)
    ws.row_dimensions[r].height = 22; r += 1
r += 1

section(ws, r, 'ビュー・図面')
r += 1
views = [
    ('平面図（真上から）',  '各ケースの配置をパレット上面から見た図。ケースごとに色分けされます。'),
    ('側面図（横から）',    'パレットを横から見た図。積み付け高さの分布を確認できます。'),
    ('等角3Dビュー',        'Three.jsによるインタラクティブな3D表示。マウスドラッグで視点を回転、スクロールでズームできます。'),
    ('アニメーション',      '▶ ボタンで積み付け順序をアニメーション再生。スライダーで任意のステップに移動できます。'),
]
for i, (name, desc) in enumerate(views):
    fl = fill(C_PALE) if i % 2 == 0 else fill(C_GRAY)
    cell(ws, r, 2, name, fnt=font(size=9, bold=True), fl=fl, align=L, border=BORDER)
    ws.merge_cells(f'C{r}:E{r}')
    cell(ws, r, 3, desc, fnt=font(size=9), fl=fl, align=LT, border=BORDER, span=3)
    ws.row_dimensions[r].height = 22; r += 1
r += 1

section(ws, r, '理想解との差分分析')
r += 1
ws.merge_cells(f'B{r}:E{r}')
ce = ws.cell(row=r, column=2,
    value='現実制約モードで計算した場合、「理想条件（制約なし）の積載率」と「実際の積載率」の差分が自動表示されます。\n'
          '差分が大きいほど、現場制約（FIFO・重量制限等）による効率ロスが大きいことを示します。')
ce.font = Font(name='メイリオ', size=9)
ce.alignment = LT; ce.fill = fill(C_PALE)
ws.row_dimensions[r].height = 40; r += 2

note_cell(ws, r,
    '💡 3Dビューの操作：マウス左ドラッグ＝視点回転 / スクロール＝ズーム / 右ドラッグ＝平行移動',
    col_span=4)

# ============================================================
# ページ11: CSV出力・注意事項
# ============================================================
ws = wb.create_sheet('9.出力と注意事項')
for i, w in enumerate([2,18,70,2], 1):
    ws.column_dimensions[get_column_letter(i)].width = w
draw_sheet_header(ws, 9, 'CSV出力・注意事項', '結果の出力方法と使用上の注意点です')

r = 4
section(ws, r, 'CSV出力', 3)
r += 1
ws.merge_cells(f'B{r}:C{r}')
ce = ws.cell(row=r, column=2,
    value='積付結果タブ（タブ④）の右上にある「CSVダウンロード」ボタンをクリックすると、'
          '積付結果をCSVファイルとして保存できます。\n'
          'ファイル名は "pallet_result.csv" として保存されます。')
ce.font = Font(name='メイリオ', size=9); ce.alignment = LT; ce.fill = fill(C_PALE)
ws.row_dimensions[r].height = 40; r += 1

ws.merge_cells(f'B{r}:C{r}')
ce = ws.cell(row=r, column=2,
    value='CSVの出力項目：品番 / 品名 / X座標 / Y座標 / Z座標 / 長さ / 幅 / 高さ / 重量 / 回転 / パレット番号 / 配置順序')
ce.font = Font(name='メイリオ', size=9); ce.alignment = LT; ce.fill = fill(C_GRAY)
ws.row_dimensions[r].height = 22; r += 2

section(ws, r, '注意事項・制限事項', 3)
r += 1
notes = [
    ('⚠', '最大品種数',       '1回の計算で登録できるケースは最大30品種です。30品種を超える場合は複数回に分けて計算してください。'),
    ('⚠', 'スリープ仕様',     '15分間アクセスがない場合、サーバーが自動的にスリープします。再アクセス時に30秒〜1分待つ場合があります。'),
    ('⚠', '計算結果の保存',   '計算結果はブラウザを閉じると消えます。必要な結果はCSVダウンロードで保存してください。'),
    ('⚠', 'ブラウザ対応',     'Internet Explorerは非対応です。Google ChromeまたはMicrosoft Edgeをご使用ください。'),
    ('⚠', '複数タブ使用',     '同じブラウザで複数タブを開いて別々の計算を行うことができます。'),
    ('💡', '設定の再現性',    '同じ入力データ・設定であれば毎回同じ計算結果が得られます（ランダムシード固定）。'),
    ('💡', '計算時間',        'ケース数が多い場合や高精度モードでは計算に数十秒かかる場合があります。'),
    ('💡', 'パレット複数枚',  '1枚に収まらない場合は自動的に2枚目・3枚目に分けて計算します（最大10枚）。'),
]
for i, (icon, name, desc) in enumerate(notes):
    fl = fill(C_LORANGE) if icon == '⚠' else fill(C_LGREEN)
    cell(ws, r, 2, f'{icon} {name}', fnt=font(size=9, bold=True), fl=fl, align=L, border=BORDER)
    cell(ws, r, 3, desc,             fnt=font(size=9),             fl=fl, align=LT, border=BORDER)
    ws.row_dimensions[r].height = 22; r += 1
r += 1

section(ws, r, 'よくある質問', 3)
r += 1
faqs = [
    ('Q. 計算結果が「未配置」になるケースがある',
     'A. ケースのサイズがパレットに対して大きすぎる、または積付制約（支持率・重量・fragile等）で配置できない場合に発生します。'
     '制約条件を緩めるか、ケースサイズを確認してください。'),
    ('Q. バッファモードとFIFOモードで結果が異なる',
     'A. 正常な動作です。供給順序の自由度が異なるため積付パターンが変わります。'
     'データ特性によってはFIFOの方が効率が高い場合もあります（考察シート参照）。'),
    ('Q. 体積効率が100%を超えた',
     'A. 体積効率は「全ケース体積÷パレット1枚分の体積」で計算されます。'
     '複数パレットを使用した場合でも分母はパレット1枚分のため、100%を超えることがあります。'),
    ('Q. 3Dビューが表示されない',
     'A. ブラウザのWebGLが無効になっている可能性があります。'
     'Google ChromeまたはEdgeの最新版をご使用ください。'),
]
for i, (q, a) in enumerate(faqs):
    fl = fill(C_PALE) if i % 2 == 0 else fill(C_GRAY)
    cell(ws, r, 2, q, fnt=font(size=9, bold=True, color=C_NAVY), fl=fl, align=LT, border=BORDER)
    cell(ws, r, 3, a, fnt=font(size=9),                           fl=fl, align=LT, border=BORDER)
    ws.row_dimensions[r].height = 40; r += 1

# ============================================================
# 保存
# ============================================================
wb.save(OUTPUT)
print(f'Done: {OUTPUT}')
print(f'シート: {wb.sheetnames}')
