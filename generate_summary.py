# -*- coding: utf-8 -*-
"""
generate_summary.py
全CSVデータに対して積み付け計算を実行し、
等角3D図付きのExcelサマリーを出力する。
"""
from __future__ import annotations
import sys
import io
import csv
import traceback
import math
from pathlib import Path

import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from mpl_toolkits.mplot3d import Axes3D  # noqa: F401
from mpl_toolkits.mplot3d.art3d import Poly3DCollection
import numpy as np

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).parent))

from models import CaseItem, PalletConfig, SupplyConfig, RuleConfig, ScoreConfig
from packer import pack


# ---------------------------------------------------------------------------
# パレット設定（UIデフォルト値に合わせる）
# ---------------------------------------------------------------------------
DEFAULT_PALLET  = PalletConfig(max_height=1800, effective_height=1650)
DEFAULT_SUPPLY  = SupplyConfig(mode='fifo', buffer_size=5)
DEFAULT_RULES   = RuleConfig(
    fragile_top=False,
    heavy_bottom=True,
    center_priority=True,
    no_overhang=False,
    full_support=False,      # 改1当時の条件を維持（部分支持 support_ratio_min=0.7 を許容）
    overhang_limit=25/1100,
    height_tolerance=5,
)
# ブラウザUIデフォルト値 (50,50,30,20,30) を正規化
_SW = 50 + 50 + 30 + 20 + 30  # 180
DEFAULT_SCORING = ScoreConfig(
    w_support = 50 / _SW,
    w_center  = 50 / _SW,
    w_height  = 30 / _SW,
    w_void    = 20 / _SW,
    w_group   = 30 / _SW,
)
# 現場再現優先プリセット（UI: stability=60, loadRate=40, height=40, void=20, grouping=70）
_FW = 60 + 40 + 40 + 20 + 70  # 230
FIELD_SCORING = ScoreConfig(
    w_support = 60 / _FW,
    w_center  = 40 / _FW,
    w_height  = 40 / _FW,
    w_void    = 20 / _FW,
    w_group   = 70 / _FW,
)
# 現場再現優先はグループ集約重み(w_group)を効かせるため same_group を有効化
from dataclasses import replace as _dc_replace
FIELD_RULES = _dc_replace(DEFAULT_RULES, same_group=True)

DEFAULT_BEAM    = 5          # 高精度（ビームサーチ width=5）
DEFAULT_EXEC    = 'real'     # 現実制約モード


# ---------------------------------------------------------------------------
# CSV 読み込み
# ---------------------------------------------------------------------------
def parse_csv(filepath: Path) -> list[CaseItem]:
    """品番,品名,長さ,幅,高さ,重量,数量 フォーマットを読む"""
    cases: list[CaseItem] = []
    with open(filepath, encoding='utf-8-sig') as f:
        reader = csv.reader(f)
        next(reader, None)          # ヘッダースキップ
        for row in reader:
            if len(row) < 7 or not row[0].strip():
                continue
            try:
                cases.append(CaseItem(
                    sku_id   = row[0].strip(),
                    name     = row[1].strip(),
                    length   = int(row[2]),
                    width    = int(row[3]),
                    height   = int(row[4]),
                    weight   = float(row[5]),
                    quantity = int(row[6]),
                ))
            except (ValueError, IndexError):
                continue
    return cases


# ---------------------------------------------------------------------------
# 等角3D図生成
# ---------------------------------------------------------------------------
SKU_COLORS = [
    '#4A90D9','#E74C3C','#2ECC71','#F39C12','#9B59B6',
    '#1ABC9C','#E67E22','#3498DB','#E91E63','#00BCD4',
    '#8BC34A','#FF5722','#607D8B','#795548','#FF9800',
    '#673AB7','#03A9F4','#8D6E63','#546E7A','#F06292',
]

def _hex_to_rgba(hex_color: str, alpha: float = 0.78) -> tuple:
    h = hex_color.lstrip('#')
    r, g, b = (int(h[i:i+2], 16) / 255 for i in (0, 2, 4))
    return (r, g, b, alpha)

def _draw_box(ax, x, y, z, dx, dy, dz, facecolor, edgecolor='#ffffff'):
    x0, x1 = x, x + dx
    y0, y1 = y, y + dy
    z0, z1 = z, z + dz
    faces = [
        [[x0,y0,z0],[x1,y0,z0],[x1,y1,z0],[x0,y1,z0]],   # bottom
        [[x0,y0,z1],[x1,y0,z1],[x1,y1,z1],[x0,y1,z1]],   # top
        [[x0,y0,z0],[x1,y0,z0],[x1,y0,z1],[x0,y0,z1]],   # front
        [[x0,y1,z0],[x1,y1,z0],[x1,y1,z1],[x0,y1,z1]],   # back
        [[x0,y0,z0],[x0,y1,z0],[x0,y1,z1],[x0,y0,z1]],   # left
        [[x1,y0,z0],[x1,y1,z0],[x1,y1,z1],[x1,y0,z1]],   # right
    ]
    poly = Poly3DCollection(faces, facecolor=facecolor,
                            edgecolor=edgecolor, linewidth=0.35)
    ax.add_collection3d(poly)

def generate_isometric_figure(placements, pallet: PalletConfig,
                               case_id: str) -> io.BytesIO:
    """ブラウザ表示に近い3D積付図を生成し BytesIO で返す"""
    fig = plt.figure(figsize=(7, 6))
    ax = fig.add_subplot(111, projection='3d')

    # ブラウザのカメラ角度に近い視点（正面やや左上から）
    ax.view_init(elev=28, azim=-60)

    # SKU → 色マッピング
    skus = list(dict.fromkeys(p.sku_id for p in placements))
    sku_rgba = {
        sku: _hex_to_rgba(SKU_COLORS[i % len(SKU_COLORS)])
        for i, sku in enumerate(skus)
    }

    # 軸設定（描画前に確定してアスペクト比に使う）
    pl, pw = pallet.length, pallet.width
    max_z = max((p.z2 for p in placements), default=pallet.effective_height)
    z_top = max(max_z, 400)

    # パレット台座
    _draw_box(ax, 0, 0, -150,
              pallet.length, pallet.width, 150,
              facecolor=(0.60, 0.50, 0.35, 0.55),
              edgecolor='#888888')

    # 各ケース描画
    for p in placements:
        _draw_box(ax, p.x, p.y, p.z,
                  p.length, p.width, p.height,
                  facecolor=sku_rgba.get(p.sku_id, (0.29, 0.56, 0.85, 0.78)))

    ax.set_xlim(0, pl)
    ax.set_ylim(0, pw)
    ax.set_zlim(-150, z_top)

    # 実寸比率でボックスアスペクトを設定（縦横比を正確に）
    ax.set_box_aspect([pl, pw, z_top + 150])

    ax.set_xlabel('X(mm)', fontsize=7, labelpad=2)
    ax.set_ylabel('Y(mm)', fontsize=7, labelpad=2)
    ax.set_zlabel('Z(mm)', fontsize=7, labelpad=2)
    ax.tick_params(axis='both', labelsize=6, pad=1)
    ax.set_title(f'Case {case_id}', fontsize=9, fontweight='bold', pad=4)

    # 凡例（最大12品種まで）
    from matplotlib.patches import Patch
    legend_skus = skus[:12]
    handles = [
        Patch(facecolor=SKU_COLORS[i % len(SKU_COLORS)], label=sku)
        for i, sku in enumerate(legend_skus)
    ]
    if handles:
        ax.legend(handles=handles, loc='upper left',
                  fontsize=6, framealpha=0.6,
                  bbox_to_anchor=(-0.05, 1.0))

    plt.tight_layout()
    buf = io.BytesIO()
    plt.savefig(buf, format='png', dpi=110, bbox_inches='tight',
                facecolor='white')
    buf.seek(0)
    plt.close(fig)
    return buf


# ---------------------------------------------------------------------------
# Excel 出力
# ---------------------------------------------------------------------------
def _make_border(style='thin'):
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)

HEADER_FILL  = PatternFill('solid', fgColor='1F4E79')
HEADER_FONT  = Font(name='メイリオ', bold=True, color='FFFFFF', size=9)
ALT_FILL     = PatternFill('solid', fgColor='EBF3FB')
CENTER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
BORDER       = _make_border()

SUMMARY_HEADERS = [
    ('ケース番号',    10),
    ('品種数',         7),
    ('総ケース数',     9),
    ('配置数',         7),
    ('未配置数',       8),
    ('パレット数',     9),
    ('体積効率(%)',   10),
    ('最大高さ(mm)',  11),
    ('総重量(kg)',    10),
    ('安定性\nスコア', 9),
    ('段数',           6),
    ('等角3D図',      42),
]

IMAGE_W_PX = 300    # Excel 画像幅 [px]
IMAGE_H_PX = 228    # Excel 画像高さ [px]
ROW_HEIGHT_PT = 174 # 行高 [pt]  ≈ IMAGE_H_PX * 72/96
MAX_PALLETS = 3     # 最大パレット数（列数）


def _fill_sheet(ws, results: list[dict]) -> None:
    """ワークシートにヘッダー＋データを書き込む"""
    pallet_img_headers = [
        (f'等角3D図({i+1}枚目)', 42) for i in range(MAX_PALLETS)
    ]
    all_headers = SUMMARY_HEADERS + pallet_img_headers

    # ---- ヘッダー行 ----
    for col, (label, width) in enumerate(all_headers, 1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border    = BORDER
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[1].height = 32

    # ---- データ行 ----
    for row_idx, d in enumerate(results, 2):
        fill = ALT_FILL if row_idx % 2 == 0 else None

        row_vals = [
            d['case_id'],
            d['sku_count'],
            d['total_cases'],
            d['placed_cases'],
            d['unplaced_cases'],
            d['pallet_count'],
            d['efficiency'],
            d['max_height'],
            d['total_weight'],
            d['stability_score'],
            d['tier_count'],
        ]
        for col, val in enumerate(row_vals, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.alignment = CENTER_ALIGN
            cell.border    = BORDER
            if fill:
                cell.fill = fill

        # パレットごとの 3D 図を各列に貼り付け
        img_bufs = d.get('img_bufs', [])
        for pallet_idx, img_buf in enumerate(img_bufs):
            if img_buf is None:
                continue
            try:
                img = XLImage(img_buf)
                img.width  = IMAGE_W_PX
                img.height = IMAGE_H_PX
                col_num = 12 + pallet_idx
                anchor = f"{get_column_letter(col_num)}{row_idx}"
                ws.add_image(img, anchor)
            except Exception as e:
                print(f"  [WARN] image insert error ({d['case_id']} pallet {pallet_idx+1}): {e}")

        ws.row_dimensions[row_idx].height = ROW_HEIGHT_PT

    ws.freeze_panes = 'A2'


def build_excel(sheet_data: list[tuple[str, list[dict]]], output_path: str) -> None:
    """
    sheet_data: [(シート名, results), ...] のリスト
    """
    wb = Workbook()
    for i, (sheet_name, results) in enumerate(sheet_data):
        ws = wb.active if i == 0 else wb.create_sheet()
        ws.title = sheet_name
        _fill_sheet(ws, results)
    wb.save(output_path)


# ---------------------------------------------------------------------------
# 計算ループ（設定を外部から指定可能）
# ---------------------------------------------------------------------------
def run_calculations(csv_files, supply: SupplyConfig, beam_width: int,
                     rules: RuleConfig = None,
                     scoring: ScoreConfig = None) -> list[dict]:
    rules   = rules if rules is not None else DEFAULT_RULES
    scoring = scoring if scoring is not None else DEFAULT_SCORING
    results: list[dict] = []
    for csv_file in csv_files:
        case_id = csv_file.stem
        print(f"[{case_id}] ", end='', flush=True)
        try:
            cases = parse_csv(csv_file)
            if not cases:
                print("  no data, skip")
                continue

            result = pack(cases, DEFAULT_PALLET, supply,
                          rules, scoring,
                          exec_mode=DEFAULT_EXEC, beam_width=beam_width)

            pallet_ids = sorted(set(p.pallet_id for p in result.placements))
            img_bufs = []
            for pid in pallet_ids:
                pallet_placements = [p for p in result.placements if p.pallet_id == pid]
                label = f"{case_id}-P{pid}" if len(pallet_ids) > 1 else case_id
                try:
                    buf = generate_isometric_figure(pallet_placements, DEFAULT_PALLET, label)
                except Exception as img_e:
                    print(f"[fig error P{pid}: {img_e}] ", end='')
                    buf = None
                img_bufs.append(buf)

            results.append({
                'case_id':        case_id,
                'sku_count':      len(cases),
                'total_cases':    result.total_cases,
                'placed_cases':   result.placed_cases,
                'unplaced_cases': len(result.unplaced),
                'pallet_count':   result.pallet_count,
                'efficiency':     result.efficiency,
                'max_height':     result.max_height_used,
                'total_weight':   result.total_weight,
                'stability_score':result.stability_score,
                'tier_count':     result.tier_count,
                'img_bufs':       img_bufs,
            })

            status = "OK" if not result.unplaced else f"NG(unplaced={len(result.unplaced)})"
            print(f"  {status}  placed={result.placed_cases}/{result.total_cases}"
                  f"  eff={result.efficiency}%  h={result.max_height_used}mm"
                  f"  w={result.total_weight}kg  pallets={result.pallet_count}")

        except Exception as e:
            print(f"ERROR: {e}")
            traceback.print_exc()
    return results


# ---------------------------------------------------------------------------
# 考察シート生成
# ---------------------------------------------------------------------------
def _load_prev_pallet_counts(path: str) -> dict:
    """改1のシートから ケース番号→{fifo/buffer: パレット数} を読み込む（比較用）"""
    from openpyxl import load_workbook
    out: dict = {}
    try:
        wb = load_workbook(path, read_only=True)
        for sheet, key in [('FIFO制約あり', 'fifo'), ('バッファあり', 'buffer')]:
            if sheet not in wb.sheetnames:
                continue
            for row in wb[sheet].iter_rows(min_row=2, values_only=True):
                if row[0] is None:
                    continue
                out.setdefault(str(row[0]), {})[key] = row[5]
        wb.close()
    except Exception as e:
        print(f"[WARN] 改1読み込み失敗（比較セクションは省略）: {e}")
    return out


def _agg(results: list[dict]) -> dict:
    n = max(len(results), 1)
    return {
        'mean_pallet':    sum(r['pallet_count'] for r in results) / n,
        'mean_eff':       sum(r['efficiency'] for r in results) / n,
        'unplaced_files': sum(1 for r in results if r['unplaced_cases'] > 0),
        'unplaced_total': sum(r['unplaced_cases'] for r in results),
    }


def build_kousatsu_rows(results_map: dict, prev: dict) -> list:
    """考察シートの行データを組み立てる。各行は (style, [col値...])"""
    rows: list = []
    add = rows.append

    add(('title', ['混載パレタイズ積付計算結果 改2  考察レポート（計算エンジン2026-06-11修正版・自動集計）']))
    add(('body',  []))

    # ---- 1. 計算条件 ----
    add(('sec',  ['1. 計算条件']))
    add(('head', ['シート', '供給モード', 'バッファ/ビーム幅', 'スコア重み', 'ルール']))
    add(('body', ['FIFO制約あり', 'FIFO（厳密順序）', 'beam=5',
                  '標準 (支持50/重心50/高さ30/空隙20/集約30)',
                  '重量物下段・中心寄せ・部分支持0.7許容・はみ出し25mm']))
    add(('body', ['バッファあり', 'バッファ(6件先読み)', 'buffer=6 / beam=6',
                  '標準 (同上)', '同上']))
    add(('body', ['FIFO制約あり_現場再現優先', 'FIFO（厳密順序）', 'beam=5',
                  '現場再現 (支持60/重心40/高さ40/空隙20/集約70)', '同上＋同品種集約ON']))
    add(('body', ['バッファあり_現場再現優先', 'バッファ(6件先読み)', 'buffer=6 / beam=6',
                  '現場再現 (同上)', '同上＋同品種集約ON']))
    add(('body', []))

    # ---- 2. 結果サマリー ----
    add(('sec',  ['2. 結果サマリー（全ケース平均）']))
    add(('head', ['シート', '平均パレット数', '平均体積効率(%)', '未配置発生ファイル数', '未配置ケース総数']))
    for name, results in results_map.items():
        a = _agg(results)
        add(('body', [name, round(a['mean_pallet'], 2), round(a['mean_eff'], 1),
                      a['unplaced_files'], a['unplaced_total']]))
    add(('body', ['※ 体積効率の定義を改2で変更: 全ケース体積 ÷ (パレット1枚分体積 × 使用パレット数)。'
                  '改1は分母がパレット1枚分固定だったため、複数パレット時の数値は改1と直接比較できない。']))
    add(('body', []))

    # ---- 3. FIFO vs バッファ ----
    fifo = {r['case_id']: r for r in results_map['FIFO制約あり']}
    buf  = {r['case_id']: r for r in results_map['バッファあり']}
    shared = [cid for cid in fifo if cid in buf]
    improved = sorted([c for c in shared if buf[c]['pallet_count'] < fifo[c]['pallet_count']])
    worsened = sorted([c for c in shared if buf[c]['pallet_count'] > fifo[c]['pallet_count']])
    add(('sec',  ['3. FIFO制約あり vs バッファあり（パレット数比較・標準重み）']))
    add(('head', ['区分', '件数', 'ケース番号']))
    add(('body', ['バッファで減少（改善）', len(improved), ', '.join(improved) or '─']))
    add(('body', ['バッファで増加（悪化）', len(worsened), ', '.join(worsened) or '─']))
    add(('body', ['同数', len(shared) - len(improved) - len(worsened), '─']))
    add(('body', []))

    # ---- 4. 標準 vs 現場再現優先 ----
    f_std = fifo
    f_fld = {r['case_id']: r for r in results_map['FIFO制約あり_現場再現優先']}
    shared2 = [cid for cid in f_std if cid in f_fld]
    fld_up   = sorted([c for c in shared2 if f_fld[c]['pallet_count'] > f_std[c]['pallet_count']])
    fld_down = sorted([c for c in shared2 if f_fld[c]['pallet_count'] < f_std[c]['pallet_count']])
    add(('sec',  ['4. 標準重み vs 現場再現優先（FIFO・パレット数比較）']))
    add(('head', ['区分', '件数', 'ケース番号']))
    add(('body', ['現場再現優先で減少', len(fld_down), ', '.join(fld_down) or '─']))
    add(('body', ['現場再現優先で増加', len(fld_up), ', '.join(fld_up) or '─']))
    add(('body', ['同数', len(shared2) - len(fld_up) - len(fld_down), '─']))
    add(('body', []))

    # ---- 5. 改1 → 改2 比較（エンジン修正の影響） ----
    if prev:
        add(('sec',  ['5. 改1 → 改2 比較（FIFO制約あり・パレット数）']))
        shared3 = [c for c in fifo if c in prev and 'fifo' in prev[c]]
        dec = sorted([c for c in shared3 if fifo[c]['pallet_count'] < prev[c]['fifo']])
        inc = sorted([c for c in shared3 if fifo[c]['pallet_count'] > prev[c]['fifo']])
        prev_mean = sum(prev[c]['fifo'] for c in shared3) / max(len(shared3), 1)
        new_mean  = sum(fifo[c]['pallet_count'] for c in shared3) / max(len(shared3), 1)
        add(('head', ['区分', '件数/値', 'ケース番号']))
        add(('body', ['平均パレット数 改1', round(prev_mean, 2), '─']))
        add(('body', ['平均パレット数 改2', round(new_mean, 2), '─']))
        add(('body', ['改2で減少', len(dec), ', '.join(dec) or '─']))
        add(('body', ['改2で増加', len(inc), ', '.join(inc) or '─']))
        add(('body', ['同数', len(shared3) - len(dec) - len(inc), '─']))
        add(('body', []))

    # ---- 6. 計算エンジン変更点 ----
    add(('sec',  ['6. 改2 計算エンジン変更点（2026-06-11）']))
    for note in [
        '・FIFO厳密化: 先頭ケースが配置不能な時点でパレットを閉じる（改1は後続を先に配置する順序違反があった）',
        '・配置失敗ケースの再試行方式を修正: 1ケースの失敗でパレットが早期に閉じる問題を解消',
        '・上面許容荷重(max_top_load)の多段積みチェック漏れを修正',
        '・温度帯分離の横隣接未検出を修正（本データでは温度帯未使用のため影響なし）',
        '・体積効率の定義変更: 使用パレット数を分母に反映（考察2参照）',
        '・1個も配置されない空パレットをパレット数に計上しないよう修正',
        '・候補生成の重複除去により計算速度を改善（結果には影響なし）',
    ]:
        add(('body', [note]))

    return rows


def _fill_kousatsu(ws, rows: list) -> None:
    title_font = Font(name='メイリオ', bold=True, size=12)
    sec_font   = Font(name='メイリオ', bold=True, size=10, color='1F4E79')
    head_font  = Font(name='メイリオ', bold=True, size=9, color='FFFFFF')
    body_font  = Font(name='メイリオ', size=9)
    widths = [4, 38, 24, 24, 40, 44]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    r = 1
    for style, vals in rows:
        for c, v in enumerate(vals, 2):
            cell = ws.cell(row=r, column=c, value=v)
            if style == 'title':
                cell.font = title_font
            elif style == 'sec':
                cell.font = sec_font
            elif style == 'head':
                cell.font = head_font
                cell.fill = HEADER_FILL
                cell.border = BORDER
                cell.alignment = CENTER_ALIGN
            else:
                cell.font = body_font
                if len(vals) > 1:
                    cell.border = BORDER
                    cell.alignment = Alignment(vertical='center', wrap_text=True)
        r += 1


# ---------------------------------------------------------------------------
# メイン
# ---------------------------------------------------------------------------
def main():
    data_dir = Path('data')
    csv_files = sorted(data_dir.glob('*.csv'), key=lambda p: p.stem)
    output_path = sys.argv[1] if len(sys.argv) > 1 else 'Summary_混載パレタイズ積付計算結果_改2.xlsx'

    print(f"CSV files found: {len(csv_files)}")

    runs = [
        ('FIFO制約あり',
         SupplyConfig(mode='fifo', buffer_size=5), DEFAULT_BEAM, DEFAULT_RULES, DEFAULT_SCORING),
        ('バッファあり',
         SupplyConfig(mode='buffer', buffer_size=6), 6, DEFAULT_RULES, DEFAULT_SCORING),
        ('FIFO制約あり_現場再現優先',
         SupplyConfig(mode='fifo', buffer_size=5), DEFAULT_BEAM, FIELD_RULES, FIELD_SCORING),
        ('バッファあり_現場再現優先',
         SupplyConfig(mode='buffer', buffer_size=6), 6, FIELD_RULES, FIELD_SCORING),
    ]

    results_map: dict = {}
    for name, supply, beam, rules, scoring in runs:
        print("\n" + "=" * 60)
        print(f"【{name}】 mode={supply.mode}, beam={beam}")
        print("=" * 60)
        results_map[name] = run_calculations(csv_files, supply, beam, rules, scoring)

    prev = _load_prev_pallet_counts('Summary_混載パレタイズ積付計算結果_改1.xlsx')

    # ---- Excel 出力（シート順は改1と同一: FIFO/バッファ/考察/現場再現×2） ----
    print("\n" + "-" * 60)
    print("Writing Excel...")
    wb = Workbook()
    ws = wb.active
    ws.title = 'FIFO制約あり'
    _fill_sheet(ws, results_map['FIFO制約あり'])
    _fill_sheet(wb.create_sheet('バッファあり'), results_map['バッファあり'])
    _fill_kousatsu(wb.create_sheet('考察'), build_kousatsu_rows(results_map, prev))
    _fill_sheet(wb.create_sheet('FIFO制約あり_現場再現優先'),
                results_map['FIFO制約あり_現場再現優先'])
    _fill_sheet(wb.create_sheet('バッファあり_現場再現優先'),
                results_map['バッファあり_現場再現優先'])
    wb.save(output_path)
    print(f"Done: {output_path}")


if __name__ == '__main__':
    main()
